#!/usr/bin/env python3
"""
실시간 인기 키워드 수집기 - 네이티브 앱
Windows/Mac 모두 지원
"""

APP_VERSION = "1.3.0"
GITHUB_REPO = "LowAHN/TrendingKeywords"

import webview
import json
import urllib.parse
import tempfile
import subprocess
import platform
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
import sys
import os

# 결과 저장 디렉터리
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(os.path.dirname(sys.executable))
else:
    BASE_DIR = Path(__file__).parent


# ─── 수집 함수 ───

import requests as req
import feedparser

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"


def _make_session():
    s = req.Session()
    s.headers.update({"User-Agent": UA})
    return s


def get_google_trending_kr():
    keywords = []
    try:
        url = "https://trends.google.co.kr/trending/rss?geo=KR"
        feed = feedparser.parse(url, request_headers={"User-Agent": UA})
        for entry in feed.entries[:20]:
            title = entry.get("title", "").strip()
            traffic = entry.get("ht_approx_traffic", "")
            if title:
                keywords.append({"keyword": title, "traffic": traffic})
    except Exception as e:
        keywords.append({"keyword": f"[오류] {e}", "traffic": ""})
    return keywords


def get_google_suggest(seed_keywords):
    session = _make_session()
    results = {}
    for seed in seed_keywords:
        try:
            encoded = urllib.parse.quote(seed)
            url = f"https://suggestqueries.google.com/complete/search?client=firefox&hl=ko&q={encoded}"
            resp = session.get(url, timeout=5)
            if resp.status_code == 200:
                results[seed] = resp.json()[1][:10]
            else:
                results[seed] = []
        except Exception:
            results[seed] = []
    return results


def get_naver_realtime():
    """Signal.bz API를 통한 네이버 실시간 검색어 Top 10"""
    session = _make_session()
    keywords = []
    try:
        url = "https://api.signal.bz/news/realtime/"
        resp = session.get(url, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            state_map = {"s": "유지", "+": "상승", "-": "하락", "n": "신규"}
            for item in data.get("top10", []):
                keywords.append({
                    "rank": item.get("rank", 0),
                    "keyword": item.get("keyword", ""),
                    "state": state_map.get(item.get("state", ""), item.get("state", "")),
                })
    except Exception as e:
        keywords.append({"rank": 0, "keyword": f"[오류] {e}", "state": ""})
    return keywords


def get_naver_suggest(seed_keywords):
    session = _make_session()
    results = {}
    for seed in seed_keywords:
        try:
            encoded = urllib.parse.quote(seed)
            url = f"https://ac.search.naver.com/nx/ac?q={encoded}&q_enc=UTF-8&st=100&frm=nv&r_format=json&r_enc=UTF-8&r_unicode=0&t_koreng=1"
            resp = session.get(url, timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                suggestions = []
                for item_group in data.get("items", [[]]):
                    for item in item_group:
                        if isinstance(item, list) and item:
                            suggestions.append(item[0])
                results[seed] = suggestions[:10]
            else:
                results[seed] = []
        except Exception:
            results[seed] = []
    return results


# ─── JS에서 호출하는 API 클래스 ───

class Api:
    def __init__(self):
        self._window = None
        self._last_result = None

    def set_window(self, window):
        self._window = window

    def get_version(self):
        return APP_VERSION

    def check_update(self):
        try:
            session = _make_session()
            resp = session.get(
                f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest",
                timeout=5,
            )
            if resp.status_code != 200:
                return {"has_update": False}
            data = resp.json()
            latest = data.get("tag_name", "").lstrip("v")
            if not latest:
                return {"has_update": False}
            has_update = self._compare_versions(latest, APP_VERSION)

            # 현재 OS에 맞는 다운로드 URL 찾기
            download_url = ""
            is_mac = platform.system() == "Darwin"
            for asset in data.get("assets", []):
                name = asset.get("name", "")
                if is_mac and name.endswith(".dmg"):
                    download_url = asset.get("browser_download_url", "")
                elif not is_mac and name.endswith(".exe"):
                    download_url = asset.get("browser_download_url", "")

            result = {
                "has_update": has_update,
                "current": APP_VERSION,
                "latest": latest,
                "url": data.get("html_url", ""),
                "download_url": download_url,
                "can_auto_update": bool(download_url) and getattr(sys, 'frozen', False),
            }
            return result
        except Exception:
            return {"has_update": False}

    def download_update(self, download_url):
        """업데이트 파일 다운로드 (진행률 콜백)"""
        try:
            session = _make_session()
            resp = session.get(download_url, stream=True, timeout=60)
            if resp.status_code != 200:
                return {"success": False, "message": "다운로드 실패"}

            total = int(resp.headers.get("content-length", 0))
            suffix = ".dmg" if platform.system() == "Darwin" else ".exe"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)

            downloaded = 0
            for chunk in resp.iter_content(chunk_size=65536):
                tmp.write(chunk)
                downloaded += len(chunk)
                if total > 0 and self._window:
                    pct = int(downloaded * 100 / total)
                    self._window.evaluate_js(f"updateProgress({pct})")

            tmp.close()
            self._update_file = tmp.name
            return {"success": True, "path": tmp.name}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def apply_update(self):
        """다운로드된 업데이트 적용 후 앱 재시작"""
        if not hasattr(self, "_update_file") or not self._update_file:
            return {"success": False, "message": "다운로드된 업데이트 없음"}

        if not getattr(sys, "frozen", False):
            return {"success": False, "message": "개발 모드에서는 자동 업데이트 불가"}

        try:
            if platform.system() == "Darwin":
                self._apply_mac_update()
            else:
                self._apply_windows_update()
            return {"success": True}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def _apply_mac_update(self):
        update_file = self._update_file
        # .app 경로 찾기: sys.executable → .app/Contents/MacOS/binary
        app_path = Path(sys.executable).parent.parent.parent
        if not str(app_path).endswith(".app"):
            raise RuntimeError(f".app 경로를 찾을 수 없음: {app_path}")

        script = f"""#!/bin/bash
sleep 2
hdiutil attach "{update_file}" -nobrowse -quiet
MOUNT=$(hdiutil info | grep "TrendingKeywords" | awk '{{print $NF}}')
if [ -d "$MOUNT/TrendingKeywords.app" ]; then
    rm -rf "{app_path}"
    cp -R "$MOUNT/TrendingKeywords.app" "{app_path}"
    hdiutil detach "$MOUNT" -quiet
    rm -f "{update_file}"
    open "{app_path}"
else
    hdiutil detach "$MOUNT" -quiet 2>/dev/null
fi
rm -f "$0"
"""
        script_path = tempfile.NamedTemporaryFile(
            delete=False, suffix=".sh", mode="w"
        )
        script_path.write(script)
        script_path.close()
        os.chmod(script_path.name, 0o755)
        subprocess.Popen(["bash", script_path.name])
        self._window.destroy()

    def _apply_windows_update(self):
        update_file = self._update_file
        current_exe = sys.executable

        script = f"""@echo off
timeout /t 3 /nobreak >nul
move /y "{update_file}" "{current_exe}"
start "" "{current_exe}"
del "%~f0"
"""
        script_path = tempfile.NamedTemporaryFile(
            delete=False, suffix=".bat", mode="w"
        )
        script_path.write(script)
        script_path.close()
        kwargs = {}
        if platform.system() == "Windows":
            kwargs["creationflags"] = 0x08000000  # CREATE_NO_WINDOW
        subprocess.Popen(["cmd", "/c", script_path.name], **kwargs)
        self._window.destroy()

    @staticmethod
    def _compare_versions(latest, current):
        try:
            lat = [int(x) for x in latest.split(".")]
            cur = [int(x) for x in current.split(".")]
            return lat > cur
        except ValueError:
            return False

    def search(self, seeds_text):
        seeds = [s.strip() for s in seeds_text.split(",") if s.strip()]
        if not seeds:
            seeds = ["오늘", "추천", "방법", "비교", "후기", "순위"]

        with ThreadPoolExecutor(max_workers=4) as pool:
            f1 = pool.submit(get_google_trending_kr)
            f2 = pool.submit(get_google_suggest, seeds)
            f3 = pool.submit(get_naver_suggest, seeds)
            f4 = pool.submit(get_naver_realtime)
            google_trends = f1.result()
            google_suggest = f2.result()
            naver_suggest = f3.result()
            naver_realtime = f4.result()

        all_keywords = []
        for item in naver_realtime:
            if item["keyword"] and not item["keyword"].startswith("[오류]"):
                all_keywords.append(item["keyword"])
        for item in google_trends:
            if item["keyword"] and not item["keyword"].startswith("[오류]"):
                all_keywords.append(item["keyword"])
        for suggestions in google_suggest.values():
            all_keywords.extend(suggestions)
        for suggestions in naver_suggest.values():
            all_keywords.extend(suggestions)
        unique = list(dict.fromkeys(all_keywords))

        result = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "naver_realtime": naver_realtime,
            "google_trends": google_trends,
            "google_suggest": google_suggest,
            "naver_suggest": naver_suggest,
            "all_unique_keywords": unique,
        }

        self._last_result = result
        return result

    def save_file(self, file_type="json"):
        if not self._last_result or not self._window:
            return {"success": False, "message": "저장할 데이터가 없습니다."}

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        data = self._last_result

        if file_type == "json":
            default_name = f"trending_{ts}.json"
            ft = ("JSON 파일 (*.json)",)
        elif file_type == "txt":
            default_name = f"trending_{ts}.txt"
            ft = ("텍스트 파일 (*.txt)",)
        elif file_type == "excel":
            default_name = f"trending_{ts}.xlsx"
            ft = ("엑셀 파일 (*.xlsx)",)
        else:
            return {"success": False, "message": "지원하지 않는 형식"}

        filepath = self._window.create_file_dialog(
            webview.SAVE_DIALOG,
            directory=str(Path.home() / "Desktop"),
            save_filename=default_name,
            file_types=ft,
        )

        if not filepath:
            return {"success": False, "message": "취소됨"}

        save_path = filepath if isinstance(filepath, str) else filepath[0]

        if file_type == "json":
            with open(save_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        elif file_type == "txt":
            lines = []
            lines.append(f"실시간 인기 키워드 리포트")
            lines.append(f"수집 시간: {data['timestamp']}")
            lines.append("=" * 50)

            lines.append(f"\n[네이버 실시간 검색어 TOP 10]")
            lines.append("-" * 40)
            for item in data.get("naver_realtime", []):
                state = f" ({item['state']})" if item.get("state") else ""
                lines.append(f"  {item['rank']:>2}. {item['keyword']}{state}")

            lines.append(f"\n[구글 연관검색어]")
            lines.append("-" * 40)
            for seed, suggestions in data["google_suggest"].items():
                lines.append(f'\n  "{seed}"')
                for s in suggestions:
                    lines.append(f"    - {s}")

            lines.append(f"\n[네이버 연관검색어]")
            lines.append("-" * 40)
            for seed, suggestions in data["naver_suggest"].items():
                lines.append(f'\n  "{seed}"')
                for s in suggestions:
                    lines.append(f"    - {s}")

            lines.append(f"\n[구글 트렌드 실시간 급상승]")
            lines.append("-" * 40)
            for i, item in enumerate(data["google_trends"], 1):
                traffic = f" ({item['traffic']})" if item.get("traffic") else ""
                lines.append(f"  {i:>2}. {item['keyword']}{traffic}")

            lines.append(f"\n{'=' * 50}")
            lines.append(f"전체 고유 키워드 ({len(data['all_unique_keywords'])}개):")
            for kw in data["all_unique_keywords"]:
                lines.append(f"  - {kw}")

            with open(save_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

        elif file_type == "excel":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()

            # 시트1: 네이버 실시간 검색어
            ws1 = wb.active
            ws1.title = "네이버 실시간"
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

            ws1.append(["순위", "키워드", "상태"])
            for cell in ws1[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for item in data.get("naver_realtime", []):
                ws1.append([item["rank"], item["keyword"], item.get("state", "")])
            ws1.column_dimensions["A"].width = 8
            ws1.column_dimensions["B"].width = 30
            ws1.column_dimensions["C"].width = 10

            # 시트2: 구글 연관검색어
            ws1b = wb.create_sheet("구글 연관검색어")
            ws1b.append(["입력 키워드", "연관 검색어"])
            for cell in ws1b[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for seed, suggestions in data["google_suggest"].items():
                for s in suggestions:
                    ws1b.append([seed, s])
            ws1b.column_dimensions["A"].width = 20
            ws1b.column_dimensions["B"].width = 40

            # 시트3: 네이버 연관검색어
            ws2 = wb.create_sheet("네이버 연관검색어")
            ws2.append(["입력 키워드", "연관 검색어"])
            for cell in ws2[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for seed, suggestions in data["naver_suggest"].items():
                for s in suggestions:
                    ws2.append([seed, s])
            ws2.column_dimensions["A"].width = 20
            ws2.column_dimensions["B"].width = 40

            # 시트4: 구글 급상승
            ws3 = wb.create_sheet("구글 급상승")
            ws3.append(["순위", "키워드", "검색량"])
            for cell in ws3[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for i, item in enumerate(data["google_trends"], 1):
                ws3.append([i, item["keyword"], item.get("traffic", "")])
            ws3.column_dimensions["A"].width = 8
            ws3.column_dimensions["B"].width = 30
            ws3.column_dimensions["C"].width = 15

            # 시트5: 전체 키워드
            ws4 = wb.create_sheet("전체 키워드")
            ws4.append(["번호", "키워드"])
            for cell in ws4[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for i, kw in enumerate(data["all_unique_keywords"], 1):
                ws4.append([i, kw])
            ws4.column_dimensions["A"].width = 8
            ws4.column_dimensions["B"].width = 40

            wb.save(save_path)

        return {"success": True, "message": str(save_path)}


# ─── HTML ───

HTML = """
<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Trending Keywords</title>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Malgun Gothic', sans-serif;
    background: #0f0f1a;
    color: #e0e0e0;
    min-height: 100vh;
    -webkit-user-select: none; user-select: none;
  }
  .container { max-width: 960px; margin: 0 auto; padding: 20px; }

  h1 {
    font-size: 26px;
    background: linear-gradient(135deg, #6ee7b7, #3b82f6);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    margin-bottom: 20px;
  }

  .input-row { display: flex; gap: 10px; margin-bottom: 15px; }
  .input-row input {
    flex: 1; padding: 12px 16px;
    background: #1a1a2e; border: 1px solid #333;
    border-radius: 10px; color: #e0e0e0; font-size: 14px;
    outline: none; transition: border 0.2s;
  }
  .input-row input:focus { border-color: #3b82f6; }

  .btn {
    padding: 10px 20px; border: none; border-radius: 10px;
    font-size: 13px; font-weight: 600; cursor: pointer;
    transition: all 0.2s;
  }
  .btn-primary { background: #3b82f6; color: white; }
  .btn-primary:hover { background: #2563eb; }
  .btn-green { background: #10b981; color: white; }
  .btn-green:hover { background: #059669; }
  .btn-purple { background: #8b5cf6; color: white; }
  .btn:disabled { opacity: 0.5; cursor: not-allowed; }

  .btn-row { display: flex; gap: 8px; margin-bottom: 15px; flex-wrap: wrap; }

  .status {
    padding: 8px 14px; background: #1a1a2e;
    border-radius: 8px; font-size: 13px; color: #6ee7b7;
    margin-bottom: 15px;
  }

  .tabs { display: flex; gap: 0; border-bottom: 2px solid #1a1a2e; }
  .tab {
    padding: 10px 18px; cursor: pointer;
    background: #1a1a2e; color: #888;
    border: none; font-size: 13px; font-weight: 500;
    border-radius: 10px 10px 0 0; transition: all 0.2s;
  }
  .tab.active { background: #1e1e32; color: #6ee7b7; }
  .tab:hover { color: #e0e0e0; }

  .panel {
    display: none; background: #1e1e32;
    border-radius: 0 10px 10px 10px; padding: 20px;
    min-height: 350px; max-height: 450px; overflow-y: auto;
  }
  .panel.active { display: block; }

  .keyword-item {
    display: flex; align-items: center; padding: 7px 0;
    border-bottom: 1px solid #252540;
  }
  .keyword-item:last-child { border-bottom: none; }
  .keyword-num { color: #555; width: 32px; font-size: 12px; text-align: right; margin-right: 12px; }
  .keyword-text { color: #6ee7b7; font-weight: 500; flex: 1; font-size: 14px; -webkit-user-select: text; user-select: text; }
  .keyword-traffic {
    color: #f59e0b; font-size: 11px;
    background: #f59e0b22; padding: 2px 8px; border-radius: 20px;
  }

  .seed-group { margin-bottom: 16px; }
  .seed-title { color: #f59e0b; font-weight: 600; font-size: 14px; margin-bottom: 6px; padding-left: 4px; }
  .suggest-item {
    padding: 4px 0 4px 20px; color: #a5b4fc; font-size: 13px;
    border-left: 2px solid #333; margin-left: 8px;
    -webkit-user-select: text; user-select: text;
  }

  .all-keyword {
    display: inline-block; background: #252540;
    padding: 5px 12px; margin: 3px; border-radius: 20px;
    font-size: 12px; color: #6ee7b7; transition: all 0.2s;
    -webkit-user-select: text; user-select: text;
  }
  .all-keyword:hover { background: #3b82f6; color: white; }

  .spinner {
    display: inline-block; width: 14px; height: 14px;
    border: 2px solid #333; border-top: 2px solid #3b82f6;
    border-radius: 50%; animation: spin 0.8s linear infinite;
    margin-right: 8px; vertical-align: middle;
  }
  @keyframes spin { to { transform: rotate(360deg); } }

  .empty { color: #555; text-align: center; padding: 40px; font-size: 14px; }

  .update-banner {
    display: none; padding: 10px 16px; margin-bottom: 15px;
    background: linear-gradient(135deg, #f59e0b22, #f59e0b11);
    border: 1px solid #f59e0b44; border-radius: 10px;
    font-size: 13px; color: #f59e0b;
    flex-direction: column; gap: 8px;
  }
  .update-banner.show { display: flex; }
  .update-top { display: flex; align-items: center; justify-content: space-between; }
  .update-top a, .update-top button {
    color: #3b82f6; text-decoration: none; font-weight: 600;
    padding: 4px 12px; background: #3b82f622; border-radius: 6px;
    border: none; cursor: pointer; font-size: 13px;
  }
  .update-top a:hover, .update-top button:hover { background: #3b82f644; }
  .update-top button:disabled { opacity: 0.5; cursor: not-allowed; }
  .progress-bar {
    display: none; height: 6px; background: #333; border-radius: 3px; overflow: hidden;
  }
  .progress-bar.show { display: block; }
  .progress-fill {
    height: 100%; width: 0%; background: linear-gradient(90deg, #3b82f6, #6ee7b7);
    border-radius: 3px; transition: width 0.2s;
  }

  .version-info {
    font-size: 11px; color: #555; margin-top: 12px; text-align: right;
  }

  ::-webkit-scrollbar { width: 6px; }
  ::-webkit-scrollbar-track { background: #1e1e32; }
  ::-webkit-scrollbar-thumb { background: #333; border-radius: 3px; }
</style>
</head>
<body>
<div class="container">
  <h1>Trending Keywords</h1>

  <div class="update-banner" id="updateBanner">
    <div class="update-top">
      <span id="updateMsg"></span>
      <span id="updateActions">
        <button id="updateBtn" onclick="doUpdate()">지금 업데이트</button>
        <a id="updateLink" href="#" target="_blank" style="display:none;margin-left:6px">수동 다운로드</a>
      </span>
    </div>
    <div class="progress-bar" id="progressBar">
      <div class="progress-fill" id="progressFill"></div>
    </div>
  </div>

  <div class="input-row">
    <input type="text" id="seeds" placeholder="키워드 입력 (쉼표 구분) - 예: 맛집, 여행, 다이어트"
           value="오늘, 추천, 방법, 비교, 후기, 순위">
  </div>

  <div class="btn-row">
    <button class="btn btn-primary" id="searchBtn" onclick="doSearch()">검색 시작</button>
    <button class="btn btn-green" id="copyBtn" onclick="copyKeywords()" disabled>키워드 복사</button>
    <button class="btn btn-purple" id="saveTxt" onclick="saveFile('txt')" disabled>TXT 저장</button>
    <button class="btn btn-purple" id="saveExcel" onclick="saveFile('excel')" disabled>엑셀 저장</button>
    <button class="btn btn-purple" id="saveJson" onclick="saveFile('json')" disabled>JSON 저장</button>
  </div>

  <div class="status" id="status">준비됨 - 키워드를 입력하고 검색을 시작하세요</div>

  <div class="tabs">
    <button class="tab active" onclick="showTab(0)">네이버 실시간</button>
    <button class="tab" onclick="showTab(1)">구글 연관검색어</button>
    <button class="tab" onclick="showTab(2)">네이버 연관검색어</button>
    <button class="tab" onclick="showTab(3)">구글 급상승</button>
    <button class="tab" onclick="showTab(4)">전체 키워드</button>
  </div>

  <div class="panel active" id="panel0"><div class="empty">검색을 시작하세요</div></div>
  <div class="panel" id="panel1"><div class="empty">검색을 시작하세요</div></div>
  <div class="panel" id="panel2"><div class="empty">검색을 시작하세요</div></div>
  <div class="panel" id="panel3"><div class="empty">검색을 시작하세요</div></div>
  <div class="panel" id="panel4"><div class="empty">검색을 시작하세요</div></div>

  <div class="version-info" id="versionInfo"></div>
</div>

<script>
let lastResult = null;
let isSearching = false;

function esc(s) {
  const d = document.createElement('div');
  d.textContent = s;
  return d.innerHTML;
}

document.getElementById('seeds').addEventListener('keydown', e => {
  if (e.key === 'Enter') doSearch();
});

let updateInfo = null;

function updateProgress(pct) {
  document.getElementById('progressFill').style.width = pct + '%';
  document.getElementById('updateMsg').textContent = '다운로드 중... ' + pct + '%';
}

async function doUpdate() {
  if (!updateInfo || !updateInfo.download_url) return;
  const btn = document.getElementById('updateBtn');
  btn.disabled = true;
  btn.textContent = '다운로드 중...';
  document.getElementById('progressBar').classList.add('show');

  try {
    const res = await window.pywebview.api.download_update(updateInfo.download_url);
    if (res && res.success) {
      document.getElementById('updateMsg').textContent = '설치 중... 앱이 재시작됩니다';
      document.getElementById('progressFill').style.width = '100%';
      btn.style.display = 'none';
      await window.pywebview.api.apply_update();
    } else {
      document.getElementById('updateMsg').textContent = '다운로드 실패: ' + (res.message || '알 수 없는 오류');
      btn.disabled = false;
      btn.textContent = '다시 시도';
    }
  } catch(e) {
    document.getElementById('updateMsg').textContent = '업데이트 오류: ' + e;
    btn.disabled = false;
    btn.textContent = '다시 시도';
  }
}

(async function checkUpdate() {
  try {
    const ver = await window.pywebview.api.get_version();
    document.getElementById('versionInfo').textContent = 'v' + ver;

    const update = await window.pywebview.api.check_update();
    if (update && update.has_update) {
      updateInfo = update;
      document.getElementById('updateMsg').textContent =
        '새 버전 v' + update.latest + ' 사용 가능 (현재 v' + update.current + ')';
      document.getElementById('updateBanner').classList.add('show');

      if (update.can_auto_update) {
        document.getElementById('updateBtn').style.display = '';
      } else {
        document.getElementById('updateBtn').style.display = 'none';
      }
      // 수동 다운로드 링크는 항상 표시
      document.getElementById('updateLink').href = update.url;
      document.getElementById('updateLink').style.display = '';
    }
  } catch(e) {}
})();

function showTab(i) {
  document.querySelectorAll('.tab').forEach((t, idx) => t.classList.toggle('active', idx === i));
  document.querySelectorAll('.panel').forEach((p, idx) => p.classList.toggle('active', idx === i));
}

async function doSearch() {
  if (isSearching) return;
  isSearching = true;

  const btn = document.getElementById('searchBtn');
  const status = document.getElementById('status');
  btn.disabled = true;
  status.innerHTML = '<span class="spinner"></span>수집 중... (10~15초 소요)';

  const seedsText = document.getElementById('seeds').value;

  try {
    const result = await window.pywebview.api.search(seedsText);
    lastResult = result;
    renderResults(result);

    document.getElementById('copyBtn').disabled = false;
    document.getElementById('saveTxt').disabled = false;
    document.getElementById('saveExcel').disabled = false;
    document.getElementById('saveJson').disabled = false;

    status.textContent = '완료! ' + result.all_unique_keywords.length + '개 키워드 수집 (' + result.timestamp + ')';
  } catch(e) {
    status.textContent = '오류: ' + e;
  }
  btn.disabled = false;
  isSearching = false;
}

function renderResults(data) {
  // 탭0: 네이버 실시간 검색어 (Signal.bz)
  let html = '';
  const stateStyle = {'상승':'color:#f6465d','하락':'color:#3b82f6','신규':'color:#f0b90b;font-weight:700','유지':'color:#848e9c'};
  const stateIcon = {'상승':'▲','하락':'▼','신규':'NEW','유지':'−'};
  if (data.naver_realtime && data.naver_realtime.length > 0) {
    data.naver_realtime.forEach(item => {
      const si = stateStyle[item.state] || 'color:#848e9c';
      const icon = stateIcon[item.state] || '';
      html += '<div class="keyword-item">' +
        '<span class="keyword-num">' + esc(String(item.rank)) + '</span>' +
        '<span class="keyword-text">' + esc(item.keyword) + '</span>' +
        '<span style="font-size:11px;padding:2px 8px;border-radius:20px;background:#ffffff08;' + si + '">' + icon + ' ' + esc(item.state) + '</span>' +
      '</div>';
    });
  } else {
    html += '<div class="empty">데이터 없음</div>';
  }
  document.getElementById('panel0').innerHTML = html;

  // 탭1: 구글 연관검색어 (입력 키워드 기반)
  html = '';
  for (const [seed, suggestions] of Object.entries(data.google_suggest)) {
    html += '<div class="seed-group"><div class="seed-title">"' + esc(seed) + '"</div>';
    suggestions.forEach(s => { html += '<div class="suggest-item">' + esc(s) + '</div>'; });
    html += '</div>';
  }
  document.getElementById('panel1').innerHTML = html || '<div class="empty">데이터 없음</div>';

  // 탭2: 네이버 연관검색어 (입력 키워드 기반)
  html = '';
  for (const [seed, suggestions] of Object.entries(data.naver_suggest)) {
    html += '<div class="seed-group"><div class="seed-title">"' + esc(seed) + '"</div>';
    suggestions.forEach(s => { html += '<div class="suggest-item">' + esc(s) + '</div>'; });
    html += '</div>';
  }
  document.getElementById('panel2').innerHTML = html || '<div class="empty">데이터 없음</div>';

  // 탭3: 구글 급상승 (한국 전체, 입력 키워드 무관)
  html = '';
  if (data.google_trends && data.google_trends.length > 0) {
    data.google_trends.forEach((item, i) => {
      html += '<div class="keyword-item">' +
        '<span class="keyword-num">' + (i+1) + '</span>' +
        '<span class="keyword-text">' + esc(item.keyword) + '</span>' +
        (item.traffic ? '<span class="keyword-traffic">' + esc(item.traffic) + '</span>' : '') +
      '</div>';
    });
  } else {
    html = '<div class="empty">데이터 없음</div>';
  }
  document.getElementById('panel3').innerHTML = html;

  // 탭4: 전체 키워드
  html = '';
  data.all_unique_keywords.forEach(kw => {
    html += '<span class="all-keyword">' + esc(kw) + '</span>';
  });
  document.getElementById('panel4').innerHTML = html || '<div class="empty">데이터 없음</div>';
}

async function saveFile(fileType) {
  if (!lastResult) return;
  const status = document.getElementById('status');
  status.textContent = '저장 위치 선택 중...';
  try {
    const res = await window.pywebview.api.save_file(fileType);
    if (res && res.success) {
      status.textContent = '저장 완료: ' + res.message;
    } else {
      status.textContent = '저장 취소됨';
    }
  } catch(e) {
    status.textContent = '저장 오류: ' + e;
  }
}

function copyKeywords() {
  if (!lastResult) return;
  const text = lastResult.all_unique_keywords.join('\\n');
  navigator.clipboard.writeText(text).then(() => {
    document.getElementById('status').textContent =
      '클립보드에 ' + lastResult.all_unique_keywords.length + '개 키워드 복사됨!';
  }).catch(() => {
    document.getElementById('status').textContent = '클립보드 복사 실패 (권한 필요)';
  });
}
</script>
</body>
</html>
"""

if __name__ == "__main__":
    api = Api()
    window = webview.create_window(
        "Trending Keywords",
        html=HTML,
        js_api=api,
        width=960,
        height=720,
        min_size=(700, 500),
        background_color="#0f0f1a",
    )
    api.set_window(window)
    webview.start()
